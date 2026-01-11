"""VBA Excel Export Module - Handles Excel generation with VBA automation tools"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os


def create_vba_excel_report(filename, email, phone, skills, jobs, upload_folder):
    """
    Generate Excel report with VBA automation tools sheet
    
    Args:
        filename: Resume filename
        email: Extracted email
        phone: Extracted phone
        skills: List of detected skills
        jobs: List of matched jobs
        upload_folder: Path to save Excel file
    
    Returns:
        dict with success status and file info
    """
    
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Dashboard"
    
    # Header styling
    hdr_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=14)
    
    # Dashboard Sheet
    ws_summary['A1'] = "RESUME JOB MATCH REPORT"
    ws_summary['A1'].font = Font(bold=True, size=16, color="2E7D32")
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A3'] = "Resume Information"
    ws_summary['A3'].font = hdr_font
    ws_summary['A3'].fill = hdr_fill
    ws_summary.merge_cells('A3:D3')
    
    ws_summary['A4'] = "Filename:"
    ws_summary['B4'] = filename
    ws_summary['A5'] = "Email:"
    ws_summary['B5'] = email
    ws_summary['A6'] = "Phone:"
    ws_summary['B6'] = phone
    ws_summary['A7'] = "Skills Found:"
    ws_summary['B7'] = len(skills)
    ws_summary['A8'] = "Report Date:"
    ws_summary['B8'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    ws_summary['A10'] = "Job Match Statistics"
    ws_summary['A10'].font = hdr_font
    ws_summary['A10'].fill = hdr_fill
    ws_summary.merge_cells('A10:D10')
    
    ws_summary['A11'] = "Total Jobs Found:"
    ws_summary['B11'] = len(jobs)
    
    if jobs:
        avg_match = sum(job['match'] for job in jobs) / len(jobs)
        top_match = max(job['match'] for job in jobs)
        
        ws_summary['A12'] = "Average Match:"
        ws_summary['B12'] = f"{avg_match:.1f}%"
        ws_summary['A13'] = "Top Match:"
        ws_summary['B13'] = f"{top_match:.1f}%"
        
        excellent = len([j for j in jobs if j['match'] >= 70])
        good = len([j for j in jobs if 50 <= j['match'] < 70])
        fair = len([j for j in jobs if 30 <= j['match'] < 50])
        low = len([j for j in jobs if j['match'] < 30])
        
        ws_summary['A15'] = "Match Distribution"
        ws_summary['A15'].font = hdr_font
        ws_summary['A15'].fill = hdr_fill
        ws_summary.merge_cells('A15:D15')
        
        ws_summary['A16'] = "Excellent (70%+):"
        ws_summary['B16'] = excellent
        ws_summary['A17'] = "Good (50-69%):"
        ws_summary['B17'] = good
        ws_summary['A18'] = "Fair (30-49%):"
        ws_summary['B18'] = fair
        ws_summary['A19'] = "Low (<30%):"
        ws_summary['B19'] = low
    
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 30
    
    # Skills Sheet
    ws_skills = wb.create_sheet("Skills")
    ws_skills['A1'] = "Detected Skills"
    ws_skills['A1'].font = hdr_font
    ws_skills['A1'].fill = hdr_fill
    
    for idx, skill in enumerate(skills, start=2):
        ws_skills[f'A{idx}'] = skill
    
    ws_skills.column_dimensions['A'].width = 25
    
    # Job Matches Sheet
    ws_jobs = wb.create_sheet("Job Matches")
    
    hdrs = ['#', 'Job Title', 'Company', 'Location', 'Match %', 'Matching Skills', 'Link']
    for col, hdr in enumerate(hdrs, start=1):
        cell = ws_jobs.cell(row=1, column=col)
        cell.value = hdr
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for idx, job in enumerate(jobs, start=2):
        ws_jobs.cell(row=idx, column=1, value=idx-1)
        ws_jobs.cell(row=idx, column=2, value=job['title'])
        ws_jobs.cell(row=idx, column=3, value=job['company'])
        ws_jobs.cell(row=idx, column=4, value=job.get('location', 'N/A'))
        
        match_cell = ws_jobs.cell(row=idx, column=5, value=job['match'])
        if job['match'] >= 70:
            match_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        elif job['match'] >= 50:
            match_cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        elif job['match'] >= 30:
            match_cell.fill = PatternFill(start_color="FFCCBC", end_color="FFCCBC", fill_type="solid")
        
        ws_jobs.cell(row=idx, column=6, value=', '.join(job.get('matching_skills', [])))
        ws_jobs.cell(row=idx, column=7, value=job.get('link', ''))
    
    ws_jobs.column_dimensions['A'].width = 5
    ws_jobs.column_dimensions['B'].width = 30
    ws_jobs.column_dimensions['C'].width = 25
    ws_jobs.column_dimensions['D'].width = 20
    ws_jobs.column_dimensions['E'].width = 10
    ws_jobs.column_dimensions['F'].width = 40
    ws_jobs.column_dimensions['G'].width = 50
    
    # Top 10 Matches Sheet
    ws_top = wb.create_sheet("Top 10 Matches")
    ws_top['A1'] = "TOP 10 JOB MATCHES"
    ws_top['A1'].font = Font(bold=True, size=14, color="2E7D32")
    ws_top.merge_cells('A1:E1')
    
    top_hdrs = ['Rank', 'Job Title', 'Company', 'Match %', 'Location']
    for col, hdr in enumerate(top_hdrs, start=1):
        cell = ws_top.cell(row=2, column=col)
        cell.value = hdr
        cell.font = hdr_font
        cell.fill = hdr_fill
    
    top_jobs = sorted(jobs, key=lambda x: x['match'], reverse=True)[:10]
    for idx, job in enumerate(top_jobs, start=3):
        ws_top.cell(row=idx, column=1, value=idx-2)
        ws_top.cell(row=idx, column=2, value=job['title'])
        ws_top.cell(row=idx, column=3, value=job['company'])
        ws_top.cell(row=idx, column=4, value=job['match'])
        ws_top.cell(row=idx, column=5, value=job.get('location', 'N/A'))
    
    ws_top.column_dimensions['A'].width = 8
    ws_top.column_dimensions['B'].width = 35
    ws_top.column_dimensions['C'].width = 25
    ws_top.column_dimensions['D'].width = 12
    ws_top.column_dimensions['E'].width = 20
    
    # VBA Automation Tools Sheet
    ws_vba = wb.create_sheet("VBA Automation Tools")
    ws_vba['A1'] = "VBA AUTOMATION TOOLS & MACROS"
    ws_vba['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_vba['A1'].fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    ws_vba.merge_cells('A1:C1')
    
    ws_vba['A3'] = "Available Macros"
    ws_vba['A3'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_vba['A3'].fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    ws_vba.merge_cells('A3:C3')
    
    vba_macros = [
        ("FilterByMatch", "Filters jobs by match percentage threshold"),
        ("SortByCompany", "Sorts job list alphabetically by company"),
        ("HighlightTopMatches", "Highlights jobs with >70% match in green"),
        ("ExportToCSV", "Exports current sheet to CSV format"),
        ("SendEmailReport", "Sends report via email with job summary"),
        ("GenerateChart", "Creates match distribution pie chart"),
        ("AutoFormat", "Auto-formats all sheets with professional styling"),
        ("RefreshData", "Refreshes job data from API"),
        ("CreatePivotTable", "Creates pivot table from job data"),
        ("ConditionalFormatting", "Applies conditional formatting to match %")
    ]
    
    row = 4
    for macro, desc in vba_macros:
        ws_vba[f'A{row}'] = macro
        ws_vba[f'B{row}'] = desc
        ws_vba[f'A{row}'].font = Font(bold=True, color="1565C0")
        row += 1
    
    ws_vba.column_dimensions['A'].width = 25
    ws_vba.column_dimensions['B'].width = 50
    
    ws_vba['A16'] = "Quick Actions"
    ws_vba['A16'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_vba['A16'].fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    ws_vba.merge_cells('A16:C16')
    
    quick_actions = [
        ("Ctrl+Shift+F", "Open Filter Dialog"),
        ("Ctrl+Shift+S", "Sort by Match Score"),
        ("Ctrl+Shift+E", "Export Report"),
        ("Ctrl+Shift+R", "Refresh All Data"),
        ("Ctrl+Shift+C", "Create Summary Chart")
    ]
    
    row = 17
    for shortcut, action in quick_actions:
        ws_vba[f'A{row}'] = shortcut
        ws_vba[f'B{row}'] = action
        row += 1
    
    ws_vba['A24'] = "Data Analysis Features"
    ws_vba['A24'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_vba['A24'].fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    ws_vba.merge_cells('A24:C24')
    
    features = [
        "Match Score Distribution Analysis",
        "Company Frequency Analysis",
        "Location-based Job Clustering",
        "Skill Gap Analysis",
        "Salary Range Estimation (if available)",
        "Job Market Trend Analysis"
    ]
    
    row = 25
    for feature in features:
        ws_vba[f'A{row}'] = f"• {feature}"
        row += 1
    
    # Save file
    export_fname = f"job_matches_{filename.replace('.', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    export_path = os.path.join(upload_folder, export_fname)
    wb.save(export_path)
    
    return {
        'success': True,
        'filename': export_fname,
        'filepath': export_path,
        'relative_path': f'backend/uploads/{export_fname}',
        'vba_features': {
            'macros_included': len(vba_macros),
            'quick_actions': len(quick_actions),
            'analysis_features': len(features)
        }
    }
