"""Resume Job Matcher - Main Flask Application"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import os
from werkzeug.utils import secure_filename
import time
from datetime import datetime
import glob

# Import custom modules
from resume_parser import extract_resume, extract_email, extract_phone, extract_skills
from job_scraper import scrape_jobs_multi, get_fallback_jobs
from job_matcher import recommend_jobs
from vba_export import create_vba_excel_report

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'docx', 'pdf', 'doc'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Global job cache
jobCache = []
cacheTime = 0
cacheDur = 1800  # 30 minutes


def allowed_file(fname):
    """Check if file extension is allowed"""
    return '.' in fname and fname.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def get_jobs_db(force_refresh=False):
    """Get jobs from cache or scrape fresh"""
    global jobCache, cacheTime
    
    cur_time = time.time()
    cache_age = cur_time - cacheTime if cacheTime > 0 else 0
    
    if not force_refresh and jobCache and cache_age < cacheDur:
        cache_min = int(cache_age / 60)
        print(f"Using cached jobs (cached {cache_min} min ago)")
        return jobCache
    
    print("Scraping fresh jobs...")
    jobs = scrape_jobs_multi(kw="software developer", loc="", max_jobs=25)
    
    scrape_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    if not jobs or len(jobs) == 0:
        print("Scraping failed, using fallback")
        jobs = get_fallback_jobs()
        for job in jobs:
            job['scraped_at'] = scrape_time
            job['is_fallback'] = True
    else:
        print(f"Got {len(jobs)} fresh jobs")
        for job in jobs:
            job['scraped_at'] = scrape_time
            job['is_fallback'] = False
    
    jobCache = jobs
    cacheTime = cur_time
    
    return jobs


# ============ API ENDPOINTS ============

@app.route('/api/upload', methods=['POST'])
def upload_resume():
    """Upload and analyze resume"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if file and allowed_file(file.filename):
        fname = secure_filename(file.filename)
        fpath = os.path.join(app.config['UPLOAD_FOLDER'], fname)
        file.save(fpath)
        
        txt = extract_resume(fpath)
        email = extract_email(txt)
        phone = extract_phone(txt)
        resume_skills = extract_skills(txt)
        
        print(f"Extracted {len(resume_skills)} skills")
        print(f"Skills: {resume_skills}")
        
        all_jobs = recommend_jobs(resume_skills, get_jobs_db(), top_n=50, min_match=0)
        
        locs = list(set([job['location'] for job in all_jobs if job['location'] != 'Not specified']))
        all_skills_set = set()
        for job in all_jobs:
            all_skills_set.update(job.get('required_skills', []))
        avail_skills = sorted(list(all_skills_set))
        
        jobs = recommend_jobs(resume_skills, get_jobs_db(), top_n=20, min_match=10)
        
        print(f"Got {len(jobs)} recommendations")
        
        if not jobs:
            print("No jobs with >10% match, lowering threshold")
            jobs = recommend_jobs(resume_skills, get_jobs_db(), top_n=20, min_match=0)
        
        cache_age = time.time() - cacheTime if cacheTime > 0 else 0
        cache_min = int(cache_age / 60)
        last_upd = datetime.fromtimestamp(cacheTime).strftime("%Y-%m-%d %H:%M:%S") if cacheTime > 0 else "Just now"
        
        return jsonify({
            'success': True,
            'filename': fname,
            'email': email,
            'phone': phone,
            'skills': resume_skills,
            'jobs': jobs,
            'total_jobs': len(jobs),
            'available_locations': locs,
            'available_skills': avail_skills[:50],
            'cache_info': {
                'last_updated': last_upd,
                'cache_age_minutes': cache_min,
                'is_fresh': cache_age < cacheDur
            }
        })
    
    return jsonify({'error': 'Invalid file type'}), 400


@app.route('/api/scrape-jobs', methods=['POST'])
def scrape_jobs():
    """Manually trigger job scraping"""
    data = request.get_json() or {}
    kw = data.get('keywords', 'software engineer')
    loc = data.get('location', '')
    max_jobs = data.get('max_jobs', 20)
    
    jobs = scrape_jobs_multi(kw=kw, loc=loc, max_jobs=max_jobs)
    
    if not jobs:
        jobs = get_fallback_jobs()
    
    global jobCache, cacheTime
    jobCache = jobs
    cacheTime = time.time()
    
    return jsonify({
        'success': True,
        'jobs_count': len(jobs),
        'jobs': jobs,
        'source': 'scraped' if jobs != get_fallback_jobs() else 'fallback'
    })


@app.route('/api/jobs', methods=['GET'])
def get_jobs():
    """Get cached jobs"""
    jobs = get_jobs_db()
    return jsonify({
        'success': True,
        'jobs_count': len(jobs),
        'jobs': jobs
    })


@app.route('/api/filter-jobs', methods=['POST'])
def filter_jobs():
    """Filter jobs by criteria"""
    data = request.get_json()
    
    resume_skills = data.get('skills', [])
    loc_filter = data.get('location')
    skill_filters = data.get('skill_filters', [])
    min_match = data.get('min_match', 10)
    
    print(f"Filtering: loc={loc_filter}, skills={skill_filters}, minMatch={min_match}")
    
    jobs = recommend_jobs(
        resume_skills, 
        get_jobs_db(),
        top_n=50, 
        min_match=min_match,
        loc_filter=loc_filter,
        skill_filter=skill_filters if skill_filters else None
    )
    
    return jsonify({
        'success': True,
        'jobs': jobs,
        'total_jobs': len(jobs)
    })


@app.route('/api/cache-status', methods=['GET'])
def cache_status():
    """Check cache status"""
    cur_time = time.time()
    cache_age = cur_time - cacheTime if cacheTime > 0 else 0
    cache_min = int(cache_age / 60)
    
    is_fresh = cache_age < cacheDur
    last_upd = datetime.fromtimestamp(cacheTime).strftime("%Y-%m-%d %H:%M:%S") if cacheTime > 0 else "Never"
    
    return jsonify({
        'cache_age_minutes': cache_min,
        'is_fresh': is_fresh,
        'last_updated': last_upd,
        'jobs_count': len(jobCache),
        'cache_duration_minutes': int(cacheDur / 60)
    })


@app.route('/api/refresh-jobs', methods=['POST'])
def refresh_jobs():
    """Force refresh jobs from live sources"""
    print("Force refreshing jobs...")
    jobs = get_jobs_db(force_refresh=True)
    
    return jsonify({
        'success': True,
        'jobs_count': len(jobs),
        'message': f'Refreshed {len(jobs)} jobs from live sources',
        'timestamp': time.time()
    })


@app.route('/api/export-excel', methods=['POST'])
def export_excel():
    """Export job matches to Excel with VBA automation tools"""
    data = request.get_json()
    
    fname = data.get('filename', 'resume')
    email = data.get('email', 'N/A')
    phone = data.get('phone', 'N/A')
    skills = data.get('skills', [])
    jobs = data.get('jobs', [])
    
    result = create_vba_excel_report(fname, email, phone, skills, jobs, app.config['UPLOAD_FOLDER'])
    
    return jsonify({
        'success': result['success'],
        'message': 'Excel report generated with VBA automation tools',
        'filename': result['filename'],
        'filepath': result['filepath'],
        'relative_path': result['relative_path'],
        'vba_features': result['vba_features']
    })


@app.route('/api/bulk-process', methods=['POST'])
def bulk_process():
    """Process multiple resumes"""
    upload_dir = app.config['UPLOAD_FOLDER']
    
    resume_files = []
    for ext in ['*.pdf', '*.docx', '*.doc']:
        resume_files.extend(glob.glob(os.path.join(upload_dir, ext)))
    
    if not resume_files:
        return jsonify({'error': 'No resume files found'}), 400
    
    results = []
    
    for fpath in resume_files:
        try:
            fname = os.path.basename(fpath)
            txt = extract_resume(fpath)
            email = extract_email(txt)
            phone = extract_phone(txt)
            skills = extract_skills(txt)
            jobs = recommend_jobs(skills, get_jobs_db(), top_n=10, min_match=10)
            
            avg_match = sum(job['match'] for job in jobs) / len(jobs) if jobs else 0
            top_match = max(job['match'] for job in jobs) if jobs else 0
            
            results.append({
                'filename': fname,
                'email': email,
                'phone': phone,
                'skills_count': len(skills),
                'skills': skills,
                'jobs_found': len(jobs),
                'avg_match': round(avg_match, 1),
                'top_match': round(top_match, 1),
                'top_jobs': jobs[:5]
            })
            
        except Exception as e:
            results.append({
                'filename': os.path.basename(fpath),
                'error': str(e)
            })
    
    # Generate bulk report Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Bulk Processing Results"
    
    hdr_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=12)
    
    hdrs = ['#', 'Filename', 'Email', 'Phone', 'Skills', 'Jobs Found', 'Avg Match %', 'Top Match %', 'Status']
    for col, hdr in enumerate(hdrs, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = hdr
        cell.font = hdr_font
        cell.fill = hdr_fill
    
    for idx, result in enumerate(results, start=2):
        ws.cell(row=idx, column=1, value=idx-1)
        ws.cell(row=idx, column=2, value=result['filename'])
        ws.cell(row=idx, column=3, value=result.get('email', 'N/A'))
        ws.cell(row=idx, column=4, value=result.get('phone', 'N/A'))
        ws.cell(row=idx, column=5, value=result.get('skills_count', 0))
        ws.cell(row=idx, column=6, value=result.get('jobs_found', 0))
        ws.cell(row=idx, column=7, value=result.get('avg_match', 0))
        ws.cell(row=idx, column=8, value=result.get('top_match', 0))
        ws.cell(row=idx, column=9, value='Error' if 'error' in result else 'Success')
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 10
    
    bulk_fname = f"bulk_processing_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    bulk_path = os.path.join(upload_dir, bulk_fname)
    wb.save(bulk_path)
    
    return jsonify({
        'success': True,
        'processed': len(results),
        'results': results,
        'excel_report': bulk_fname
    })


@app.route('/api/download-bulk-report/<fname>', methods=['GET'])
def download_bulk_report(fname):
    """Download bulk processing report"""
    fpath = os.path.join(app.config['UPLOAD_FOLDER'], fname)
    if os.path.exists(fpath):
        return send_file(fpath, as_attachment=True, download_name=fname)
    return jsonify({'error': 'File not found'}), 404


@app.route('/api/health', methods=['GET'])
def health():
    """Health check"""
    return jsonify({'status': 'ok'})


if __name__ == '__main__':
    app.run(debug=True, port=5000)
